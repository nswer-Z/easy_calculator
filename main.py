from tkinter import *
from tkinter import ttk
import pandas as pd
from pandas import *
from numpy import *
from openpyxl import *

# datalast = read_excel('answer.xlsx')
# datalast = datalast.iloc[:, 1:3]
# print(datalast)
result = []  # 储存计算结果
formula = []  # 储存计算式
# 初次赋值上次运行历史记录，之后变为上次运行历史+本次运行历史
formula_last = []
result_last = []


class Empty:
    def get_children(self):
        return [0]

    def delete(self, m):
        pass

    def insert(self, a, s, values):
        pass


sm = Empty()


class Calculator:

    def __init__(self, master):
        self.master = master
        self.master.title("Calculator")
        self.master.resizable(0, 0)  # 设置窗口不可拉伸

        # 主窗体居中屏幕
        screenwidth = self.master.winfo_screenwidth()
        screenheight = self.master.winfo_screenheight()
        size = '%dx%d+%d+%d' % (800, 420, (screenwidth - 800) / 2, (screenheight - 420) / 2)
        self.master.geometry(size)  # 设置主窗口的初始尺寸 宽800，高420
        self.master.update()

        self.result = StringVar()  # 用于显示结果的可变文本
        self.equation = StringVar()  # 显示计算方程
        self.result.set(' ')
        self.equation.set('0')
        # 显示框
        self.show_result_eq = Label(self.master, bg='white', fg='black',
                                    font=('Arail', '16'), bd='0',
                                    textvariable=self.equation, anchor='se')
        self.show_result = Label(self.master, bg='white', fg='black',
                                 font=('Arail', '20'), bd='0',
                                 textvariable=self.result, anchor='se')
        # 按钮
        self.button_back = Button(self.master, text='<-', bg='DarkGray', command=self.back)  # 返回
        self.button_lbracket = Button(self.master, text='(', bg='DarkGray', command=lambda: self.getNum('('))  # 左括号
        self.button_rbracket = Button(self.master, text=')', bg='DarkGray', command=lambda: self.getNum(')'))  # 左括号
        self.button_division = Button(self.master, text='÷', bg='DarkGray', command=lambda: self.getNum('÷'))  # 除号
        # 7 8 9 4 5 6 1 2 3
        self.button_7 = Button(self.master, text='7', bg='DarkGray', command=lambda: self.getNum('7'))  # 7号
        self.button_8 = Button(self.master, text='8', bg='DarkGray', command=lambda: self.getNum('8'))  # 8号
        self.button_9 = Button(self.master, text='9', bg='DarkGray', command=lambda: self.getNum('9'))  # 9号
        self.button_multiplication = Button(self.master, text='*', bg='DarkGray',
                                            command=lambda: self.getNum('*'))  # 乘号
        # 按钮的command参数,是回调函数。lambda函数是为了可以传参数给回调函数
        self.button_4 = Button(self.master, text='4', bg='DarkGray', command=lambda: self.getNum('4'))  # 4号
        self.button_5 = Button(self.master, text='5', bg='DarkGray', command=lambda: self.getNum('5'))  # 5号
        self.button_6 = Button(self.master, text='6', bg='DarkGray', command=lambda: self.getNum('6'))  # 6号
        self.button_minus = Button(self.master, text='-', bg='DarkGray', command=lambda: self.getNum('-'))  # -号

        self.button_1 = Button(self.master, text='1', bg='DarkGray', command=lambda: self.getNum('1'))  # 1号
        self.button_2 = Button(self.master, text='2', bg='DarkGray', command=lambda: self.getNum('2'))  # 2号
        self.button_3 = Button(self.master, text='3', bg='DarkGray', command=lambda: self.getNum('3'))  # 3号
        self.button_plus = Button(self.master, text='+', bg='DarkGray', command=lambda: self.getNum('+'))  # +号
        # 控制按钮 0 .
        self.button_MC = Button(self.master, text='MC', bg='DarkGray', command=self.clear)  # MC
        self.button_0 = Button(self.master, text='0', bg='DarkGray', command=lambda: self.getNum('0'))  # 0
        self.button_dot = Button(self.master, text='.', bg='DarkGray', command=lambda: self.getNum('.'))  # .
        self.button_eq = Button(self.master, text='=', bg='DarkGray', command=self.run)  # =

        # Layout布局
        self.show_result_eq.place(x='10', y='10', width='300', height='50')
        self.show_result.place(x='10', y='60', width='300', height='50')

        self.button_back.place(x='10', y='150', width='60', height='40')
        self.button_lbracket.place(x='90', y='150', width='60', height='40')
        self.button_rbracket.place(x='170', y='150', width='60', height='40')
        self.button_division.place(x='250', y='150', width='60', height='40')
        self.button_7.place(x='10', y='205', width='60', height='40')
        self.button_8.place(x='90', y='205', width='60', height='40')
        self.button_9.place(x='170', y='205', width='60', height='40')
        self.button_multiplication.place(x='250', y='205', width='60', height='40')

        self.button_4.place(x='10', y='260', width='60', height='40')
        self.button_5.place(x='90', y='260', width='60', height='40')
        self.button_6.place(x='170', y='260', width='60', height='40')
        self.button_minus.place(x='250', y='260', width='60', height='40')

        self.button_1.place(x='10', y='315', width='60', height='40')
        self.button_2.place(x='90', y='315', width='60', height='40')
        self.button_3.place(x='170', y='315', width='60', height='40')
        self.button_plus.place(x='250', y='315', width='60', height='40')

        self.button_MC.place(x='10', y='370', width='60', height='40')
        self.button_0.place(x='90', y='370', width='60', height='40')
        self.button_dot.place(x='170', y='370', width='60', height='40')
        self.button_eq.place(x='250', y='370', width='60', height='40')

        # 查询表格
        columns = ("结果", "计算式")
        self.master = ttk.Treeview(self.master, height=18, show="headings", columns=columns)  # 表格
        global sm
        sm = self.master
        self.master.column("结果", minwidth=150, anchor='w')  # 表示列,不显示
        self.master.column("计算式", width=500, anchor='w')
        self.master.heading("结果", text="结果", anchor='w')  # 显示表头
        self.master.heading("计算式", text="计算式", anchor='w')
        self.master.pack(side=RIGHT, fill=BOTH)
        self.master.place(x='350', y='10', width='400', height='400')  # 表格位置
        # 水平滚轮
        self.VScroll2 = Scrollbar(self.master, orient='horizontal', command=self.master.xview)
        self.VScroll2.place(relx=0.000, rely=0.954, relwidth=1.000, relheight=0.046)
        self.master.configure(xscrollcommand=self.VScroll2.set)

        # 显示上一次运行程序的计算历史
        self.display_last()
        for k in range(min(len(formula_last), len(result_last))):  # 写入数据
            self.master.insert('', k, values=(formula_last[k], result_last[k]))

    def back(self):
        temp_equ = self.equation.get()
        self.equation.set(temp_equ[:-1])  # 一个一个删

    def getNum(self, arg):
        temp_equ = self.equation.get()  # 输入算式
        temp_result = self.result.get()

        # 判断基本语法错误
        if temp_result != ' ':  # 计算器输入前还没有结果，那么结果区域应该设置为空。
            self.result.set(' ')
        if temp_equ == '0' and (arg not in ['.', '+', '-', '*', '÷']):  # 如果首次输入为0，则紧跟则不能是数字，只是小数点或运算符
            temp_equ = ''
        if len(temp_equ) > 2 and temp_equ[-1] == '0':  # 运算符后面也不能出现0+数字的情形03，09，x
            if (temp_equ[-2] in ['+', '-', '*', '÷']) and (
                    arg in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '(']):
                temp_equ = temp_equ[:-1]
        temp_equ = temp_equ + arg
        self.equation.set(temp_equ)

    def clear(self):
        self.equation.set('0')
        self.result.set(' ')

    temp_equ = ''
    answer = ''

    def run(self):
        temp_equ = self.equation.get()
        temp_equ = temp_equ.replace('÷', '/')
        if temp_equ[0] in ['+', '-', '*', '÷']:
            temp_equ = '0' + temp_equ
            print(temp_equ)
        try:
            answer = '%.4f' % eval(temp_equ)  # 保留两位小数
            formula.append(temp_equ)
            formula_last.append(temp_equ)
            self.result.set(str(answer))
            result.append(answer)
            result_last.append(answer)
            print(result_last)
            print(formula_last)
            # 清除上次显示
            x = sm.get_children()
            for item in x:
                sm.delete(item)
            # 刷新本次显示，相对上次显示增加本次计算式和结果
            for k in range(min(len(formula_last), len(result_last))):  # 写入数据
                sm.insert('', k, values=(formula_last[k], result_last[k]))
            df = pd.DataFrame({"计算式": formula, "结果": result})
            print(df)
            df.to_excel(r'.\answer.xlsx')
        except (ZeroDivisionError, SyntaxError):  # 其他除0错误，或语法错误返回Error
            self.result.set(str('Error'))

    def display_last(self):
        self.workbook = load_workbook(r'.\answer.xlsx')
        sheet1 = self.workbook['Sheet1']
        i = 2
        while str(sheet1.cell(row=i, column=2).value) != 'None':
            formula_last.append(str(sheet1.cell(row=i, column=2).value))
            i += 1
        j = 2
        while str(sheet1.cell(row=j, column=3).value) != 'None':
            result_last.append(str(sheet1.cell(row=j, column=3).value))
            j += 1


if __name__ == "__main__":
    root = Tk()
    my_cal = Calculator(root)
    root.mainloop()
