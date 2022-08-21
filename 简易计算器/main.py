from tkinter import *
from tkinter import ttk
import pandas as pd
from pandas import *
from numpy import *
from openpyxl import *
import math

result = []  # 储存本次运行计算结果
formula = []  # 储存本次运行计算式

# 初次赋值上次运行历史记录，之后变为上次运行历史+本次运行历史
formula_last = []
result_last = []

b = False  # 判断是否按下'='，如果是，则清空上一次计算式，如果否，则不清空


class Calculator:

    def __init__(self, master):
        self.master = master
        self.master.title("Calculator")
        self.master.resizable(0, 0)  # 设置窗口不可拉伸

        # 主窗体居中屏幕
        screenwidth = self.master.winfo_screenwidth()
        screenheight = self.master.winfo_screenheight()
        size = '%dx%d+%d+%d' % (760, 490, (screenwidth - 800) / 2, (screenheight - 420) / 2)
        self.master.geometry(size)  # 设置主窗口的初始尺寸 宽760，高420
        self.master.update()

        self.result = StringVar()  # 用于显示结果的可变文本
        self.equation = StringVar()  # 显示计算方程
        self.result.set(' ')
        self.equation.set('0')
        # 显示框
        self.show_result_eq = Label(self.master, bg='white', fg='black',
                                    font=('Arail', '16'), bd='0',
                                    textvariable=self.equation, anchor='se')
        self.show_result = Label(self.master, bg='white', fg='black',
                                 font=('Arail', '20'), bd='0',
                                 textvariable=self.result, anchor='se')
        # 按钮
        # 第一行按钮 'back'、'('、')'、'÷'
        self.button_back = Button(self.master, text='<-', bg='DarkGray', command=self.back)  # 返回
        self.button_lbracket = Button(self.master, text='(', bg='DarkGray', command=lambda: self.getNum('('))  # 左括号
        self.button_rbracket = Button(self.master, text=')', bg='DarkGray', command=lambda: self.getNum(')'))  # 左括号
        self.button_division = Button(self.master, text='÷', bg='DarkGray', command=lambda: self.getNum('÷'))  # 除号
        # 第二行按钮 '7'、'8'、'9'、'*'
        self.button_7 = Button(self.master, text='7', bg='DarkGray', command=lambda: self.getNum('7'))  # 7号
        self.button_8 = Button(self.master, text='8', bg='DarkGray', command=lambda: self.getNum('8'))  # 8号
        self.button_9 = Button(self.master, text='9', bg='DarkGray', command=lambda: self.getNum('9'))  # 9号
        self.button_multiplication = Button(self.master, text='*', bg='DarkGray',
                                            command=lambda: self.getNum('*'))  # 乘号
        # 第三行按钮 '4'、'5'、'6'、'-'
        self.button_4 = Button(self.master, text='4', bg='DarkGray', command=lambda: self.getNum('4'))  # 4号
        self.button_5 = Button(self.master, text='5', bg='DarkGray', command=lambda: self.getNum('5'))  # 5号
        self.button_6 = Button(self.master, text='6', bg='DarkGray', command=lambda: self.getNum('6'))  # 6号
        self.button_minus = Button(self.master, text='-', bg='DarkGray', command=lambda: self.getNum('-'))  # -号
        # 第四行按钮 '1'、'2'、'3'、'+'
        self.button_1 = Button(self.master, text='1', bg='DarkGray', command=lambda: self.getNum('1'))  # 1号
        self.button_2 = Button(self.master, text='2', bg='DarkGray', command=lambda: self.getNum('2'))  # 2号
        self.button_3 = Button(self.master, text='3', bg='DarkGray', command=lambda: self.getNum('3'))  # 3号
        self.button_plus = Button(self.master, text='+', bg='DarkGray', command=lambda: self.getNum('+'))  # +号
        # 第五行按钮 'log'、'0'、'.'、'pi'
        self.button_log = Button(self.master, text='ln', bg='DarkGray', command=lambda: self.getNum('log'))  # log
        self.button_0 = Button(self.master, text='0', bg='DarkGray', command=lambda: self.getNum('0'))  # 0
        self.button_dot = Button(self.master, text='.', bg='DarkGray', command=lambda: self.getNum('.'))  # .
        self.button_pi = Button(self.master, text='π', bg='DarkGray', command=lambda: self.getNum('π'))  # pi
        # 第六行按钮 'MC'、'exp'、'e'、'='
        self.button_MC = Button(self.master, text='MC', bg='DarkGray', command=self.clear)  # MC
        self.button_exp = Button(self.master, text='exp', bg='DarkGray', command=lambda: self.getNum('exp'))  # exp
        self.button_e = Button(self.master, text='e', bg='DarkGray', command=lambda: self.getNum('e'))  # e
        self.button_eq = Button(self.master, text='=', bg='DarkGray', command=self.run)  # =

        # 按钮属性设置，包括位置、大小
        # 结果和计算式标签
        self.show_result_eq.place(x='10', y='10', width='300', height='50')
        self.show_result.place(x='10', y='60', width='300', height='50')
        # 第一行按钮 'back'、'('、')'、'÷'
        self.button_back.place(x='10', y='150', width='60', height='40')
        self.button_lbracket.place(x='90', y='150', width='60', height='40')
        self.button_rbracket.place(x='170', y='150', width='60', height='40')
        self.button_division.place(x='250', y='150', width='60', height='40')
        # 第二行按钮 '7'、'8'、'9'、'*'
        self.button_7.place(x='10', y='205', width='60', height='40')
        self.button_8.place(x='90', y='205', width='60', height='40')
        self.button_9.place(x='170', y='205', width='60', height='40')
        self.button_multiplication.place(x='250', y='205', width='60', height='40')
        # 第三行按钮 '4'、'5'、'6'、'-'
        self.button_4.place(x='10', y='260', width='60', height='40')
        self.button_5.place(x='90', y='260', width='60', height='40')
        self.button_6.place(x='170', y='260', width='60', height='40')
        self.button_minus.place(x='250', y='260', width='60', height='40')
        # 第四行按钮 '1'、'2'、'3'、'+'
        self.button_1.place(x='10', y='315', width='60', height='40')
        self.button_2.place(x='90', y='315', width='60', height='40')
        self.button_3.place(x='170', y='315', width='60', height='40')
        self.button_plus.place(x='250', y='315', width='60', height='40')
        # 第五行按钮 'log'、'0'、'.'、'pi'
        self.button_log.place(x='10', y='370', width='60', height='40')
        self.button_0.place(x='90', y='370', width='60', height='40')
        self.button_dot.place(x='170', y='370', width='60', height='40')
        self.button_pi.place(x='250', y='370', width='60', height='40')
        # 第六行按钮 'MC'、'exp'、'e'、'='
        self.button_MC.place(x='10', y='425', width='60', height='40')
        self.button_exp.place(x='90', y='425', width='60', height='40')
        self.button_e.place(x='170', y='425', width='60', height='40')
        self.button_eq.place(x='250', y='425', width='60', height='40')
        # 历史记录表格
        columns = ("结果", "计算式")
        self.master = ttk.Treeview(self.master, height=18, show="headings", columns=columns)  # 表格
        self.master.column("结果", width=150, anchor='w')  # 表格列设置
        self.master.column("计算式", width=500, anchor='w')
        self.master.heading("结果", text="结果", anchor='w', )  # 表头设置
        self.master.heading("计算式", text="计算式", anchor='w')
        self.master.pack(side=RIGHT, fill=BOTH)
        self.master.place(x='350', y='10', width='400', height='460')  # 表格位置
        # 底部水平滚轮
        self.VScroll2 = Scrollbar(self.master, orient='horizontal', command=self.master.xview)
        self.VScroll2.place(relx=0.000, rely=0.954, relwidth=1.000, relheight=0.046)
        self.master.configure(xscrollcommand=self.VScroll2.set)

        # 显示上一次运行程序的计算历史
        self.display_last()
        for k in range(min(len(formula_last), len(result_last))):  # 写入数据
            self.master.insert('', k, values=(result_last[k], formula_last[k]))

    def back(self):
        temp_equ = self.equation.get()
        self.equation.set(temp_equ[:-1])  # 一个一个删

    def getNum(self, arg):
        # 当上一次按下'=' 时，本次先清除上一次的计算式
        global b
        if b:
            self.equation_clear()
            b = False
        temp_equ = self.equation.get()  # 输入算式
        temp_result = self.result.get()
        arg, temp_equ = self.equtionGenerateError(arg, temp_equ, temp_result)
        temp_equ = temp_equ + arg
        if arg in ['log', 'exp']:  # math.log()和math.exp()的运算需要加括号，右括号使用者会自行补齐
            temp_equ = temp_equ + '('
        self.equation.set(temp_equ)

    # 判断基本语法错误
    def equtionGenerateError(self, arg, temp_equ, temp_result):
        if (temp_equ == '' or temp_equ == '0') and (arg in ['e', 'π', 'log', 'exp']):  # 第一个输入非数字时被允许的情况
            pass
        elif temp_equ == '' or temp_equ == '0' and arg in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '(']:
            pass
        elif (temp_equ[-1] in ['exp', 'log', 'π', 'e', '.']) and (arg in ['exp', 'log', 'π', 'e', '.']):  # 不符合四则运算的运算符号连接
            arg = ''
        elif (temp_equ[-1] in ['+', '-', '*', '÷', '(', '.']) and (arg in ['+', '-', '*', '÷', ')', '.']):
            arg = ''
        elif temp_equ[-1] == '.' and arg == '(':
            arg = ''
        elif temp_equ[-1] == ')' and arg == '.':
            arg = ''
        elif temp_equ[-1] == ')' and arg not in ['+', '-', '*', '÷', ')']:  # ')'后目前只能跟'+', '-', '*', '÷', ')'
            arg = ''
        elif (arg == ')') and ('(' not in temp_equ):  # 必须有左括号才能输右括号
            arg = ''
        elif arg == ')':
            lbracket_num = temp_equ.count('(')
            rbracket_num = temp_equ.count(')')
            if rbracket_num == lbracket_num:  # 左右括号的数量必须相等
                arg = ''
        elif (temp_equ[-1] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'π', 'e']) and (
                arg in ['(', 'exp', 'log', 'π', 'e']):  # 不能实现点乘
            arg = ''
        elif (temp_equ[-1] in ['π', 'e']) and (
                arg in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'π', 'e']):
            arg = ''
        elif len(temp_equ) > 1 and temp_equ[-2] == '(' and temp_equ[-1] == '0' and arg != '.':    # '('后不能跟'0'
            arg = ''
        if arg == '.' and temp_equ.count('.') > 0:  # 一个数之中只能有一个小数点
            i = -1
            while True:
                if temp_equ[i] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
                    i -= 1
                    continue
                elif temp_equ[i] == '.':
                    arg = ''
                    break
                else:
                    break
        if temp_result != ' ':  # 计算器输入前还没有结果，那么结果区域应该设置为空。
            self.result.set(' ')
        if temp_equ == '0' and (arg not in ['.', '+', '-', '*', '÷']):  # 如果首次输入为0，则紧跟则不能是数字，只是小数点或运算符
            temp_equ = ''
        if len(temp_equ) > 2 and temp_equ[-1] == '0':  # 运算符后面也不能出现0+数字的情形03，09，x
            if (temp_equ[-2] in ['+', '-', '*', '÷']) and (
                    arg in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '(']):
                temp_equ = temp_equ[:-1]
        return arg, temp_equ

    # 计算式和计算结果标签的清零、清空
    def clear(self):
        self.equation.set('0')
        self.result.set(' ')

    def equation_clear(self):
        self.equation.set('0')

    def result_clear(self):
        self.result.set(' ')

    temp_equ = ''
    answer = ''

    def run(self):
        global b
        b = True  # 点击'='调用此方法，使得b为True，下次点击按钮通过getNum()方法可以使得计算式标签清零

        temp_equ = self.equation.get()
        temp_equ = temp_equ.replace('π', 'pi')
        temp_equ = temp_equ.replace('÷', '/')
        if temp_equ[0] in ['+', '-', '*', '÷']:
            temp_equ = '0' + temp_equ
        try:
            answer = '%.4f' % eval(temp_equ)  # 保留四位小数
            self.result.set(str(answer))
            temp_equ = temp_equ.replace('pi', 'π')
            formula.append(temp_equ)  # 储存本次记录
            result.append(answer)
            formula_last.append(temp_equ)  # 将本次记录添加至上次历史记录，一起显示
            result_last.append(answer)
            # 清除上次显示
            x = self.master.get_children()
            for item in x:
                self.master.delete(item)
            # 写入数据，刷新本次显示，相对上次显示增加本次计算式和结果
            for k in range(min(len(formula_last), len(result_last))):  #
                self.master.insert('', k, values=(result_last[k], formula_last[k]))
            # 将本次记录写入answer.xlsx储存
            df = pd.DataFrame({"结果": result, "计算式": formula})
            df.to_excel(r'.\answer.xlsx')
        except:  # 其他除0错误，或语法错误返回Error
            self.result.set(str('Error'))

    # 读取answer.xlsx中的历史记录
    def display_last(self):
        self.workbook = load_workbook(r'.\answer.xlsx')
        sheet1 = self.workbook['Sheet1']
        i = 2
        while str(sheet1.cell(row=i, column=3).value) != 'None':
            formula_last.append(str(sheet1.cell(row=i, column=3).value))
            i += 1
        j = 2
        while str(sheet1.cell(row=j, column=2).value) != 'None':
            result_last.append(str(sheet1.cell(row=j, column=2).value))
            j += 1


if __name__ == "__main__":
    root = Tk()
    my_cal = Calculator(root)
    root.mainloop()
